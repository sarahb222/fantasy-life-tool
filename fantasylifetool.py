# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import *
from shutil import copyfile
import pandas as pd
from openpyxl import load_workbook
import sys
import webbrowser
import array

URLCol = 3
QNameCol = 7
TurnInCol = 9
LocationFirstCol=10
LocationLastCol=16

#0=unobtained, 1=obtained, 2=completed, 3=turned in, 4=all
TopButton=4



class Location():
    def __init__(self, widget):
        self.widget = widget
        self.photo = None
        self.label = None
        self.text = tk.StringVar()
        self.b = None
        self.key = None
        self.array = array.array('i',[0,0,0,0])
    

class ToolTip(object):

    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0

    def showtip(self, text):
        "Display text in tooltip window"
        self.text = text
        if self.tipwindow or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 27
        y = y + cy + self.widget.winfo_rooty() +27
        self.tipwindow = tw = Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry("+%d+%d" % (x, y))
        try:
            # For Mac OS
            tw.tk.call("::tk::unsupported::MacWindowStyle",
                       "style", tw._w,
                       "help", "noActivates")
        except TclError:
            pass
        label = Label(tw, text=self.text, justify=LEFT,
                      background="#ffffe0", relief=SOLID, borderwidth=1,
                      font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()
            

class Scrollbar(tk.Frame):
    def __init__(self, root):

        tk.Frame.__init__(self, root)
        self.canvas = tk.Canvas(root, borderwidth=0)
        self.frame = tk.Frame(self.canvas)
        self.vsb = tk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)
        self.hsb = tk.Scrollbar(root, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(xscrollcommand=self.hsb.set)
    
        self.vsb.pack(side="right", fill="y")
        self.hsb.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.create_window((4,4), window=self.frame, anchor="nw", 
                                  tags="self.frame")
        
        self.frame.bind("<Configure>", self.onFrameConfigure)
        
        self.minrange=2
        self.maxrange=30
        self.count = 2
        self.finish = 0
        self.noforward = 0
        self.prevlist = []
        self.abssmall = 0
    
        self.populate()
    
    
    def setText(obj):
        obj.text.set(str(obj.array[0]) + " / " + str(obj.array[1]) + " / "  + str(obj.array[2]) + " / "  + str(obj.array[3]))
        
    def nameToObj(name):
        if(name=="Castele Castle"):
            global cc
            return cc
        elif(name=="Castele Square"):
            global cs
            return cs
        elif(name=="Castele Shopping District"):
            global csd
            return csd
        elif(name=="Castele Artisan District"):
            global cad
            return cad
        elif(name=="Castele Outskirts"):
            global co
            return co
        elif(name=="South Castele"):
            global sc
            return sc
        elif(name=="West Castele"):
            global wc
            return wc
        elif(name=="East Castele"):
            global ec
            return ec
        elif(name=="East Grassy Plains"):
            global egp
            return egp
        elif(name=="Haniwa Cave"):
            global hc
            return hc
        elif(name=="Mount Snowpeak"):
            global ms
            return ms
        elif(name=="Lava Cave"):
            global lc
            return lc
        elif(name=="Waterfall Cave"):
            global wfc
            return wfc
        elif(name=="Mount Snowpeak Summit"):
            global mss
            return mss
        elif(name=="Elderwood"):
            global ew
            return ew
        elif(name=="Deep Elderwood"):
            global de
            return de
        elif(name=="Elderwood Village"):
            global ewv
            return ewv
        elif(name=="Spirit Tree"):
            global st
            return st
        elif(name=="West Grassy Plains"):
            global wgp
            return wgp
        elif(name=="Farley's Plantation"):
            global fp
            return fp
        elif(name=="Desert Ravine"):
            global dr
            return dr
        elif(name=="Port Puerto Palace"):
            global ppp
            return ppp
        elif(name=="Port Puerto Palace Way"):
            global pppw
            return pppw
        elif(name=="Port Puerto Marina"):
            global ppm
            return ppm
        elif(name=="Port Puerto Beach District"):
            global ppbd
            return ppbd
        elif(name=="Tortuga Archipelago"):
            global ta
            return ta
        elif(name=="Nautilus Cave"):
            global nc
            return nc
        elif(name=="Deepsea Cave"):
            global dc
            return dc
        elif(name=="Dark Sultan's Fortress"):
            global dsf
            return dsf
        elif(name=="Al Maajik Spelltown"):
            global amsp
            return amsp
        elif(name=="Al Maajik Sandtown"):
            global amsa
            return amsa
        elif(name=="Al Maajik Outskirts"):
            global amo
            return amo
        elif(name=="Aridian Desert"):
            global ad
            return ad
        elif(name=="Cave of Bones"):
            global cb
            return cb
        elif(name=="Subterranean Lake"):
            global sl
            return sl
        elif(name=="Ancient Ruins"):
            global ar
            return ar
        elif(name=="Cacto Cove"):
            global cac
            return cac
        elif(name=="Cave of Shadows"):
            global cos
            return cos
        elif(name=="Plushling Camp"):
            global pc
            return pc
        elif(name=="Terra Nimbus"):
            global tn
            return tn
        elif(name=="Central Grassland"):
            global cg
            return cg
        elif(name=="Forest Shrine"):
            global fs
            return fs
        elif(name=="Rocky Hill Shrine"):
            global rhs
            return rhs
        elif(name=="Penguin Beach"):
            global pb
            return pb
        elif(name=="Furlin's Grotto"):
            global fg
            return fg
        elif(name=="Ancient Tower"):
            global at
            return at
        
    def findCol(obj):
        j = TurnInCol
        while(wb['Sheet1'].cell(row=1, column=j).value != None):
            if wb['Sheet1'].cell(row=1, column=j).value == obj.key:
                return j
            j=j+1
        return -1
        
    def populate(self):
        
        tkvar = []
        def goForward(*args):
            if ((self.maxrange - self.minrange) == 28) & (self.noforward==0):
                self.finish=0
                if(self.prevlist == []):
                    self.abssmall = self.minrange
                self.prevlist.append(self.minrange)
                self.minrange=self.maxrange+1
                self.maxrange=self.minrange+28
                self.count=2
                self.frame.destroy()
                self.frame = tk.Frame(self.canvas)
                self.frame.bind("<Configure>", self.onFrameConfigure)
                self.canvas.create_window((4,4), window=self.frame, anchor="nw", 
                                      tags="self.frame")
                self.populate()
                
        def goBack(*args):
            if(self.prevlist == []):
                if (self.abssmall != 0):
                    self.minrange = self.abssmall
                    self.maxrange = self.minrange + 28
                else:
                    self.minrange = 2
                    self.maxrange = self.minrange + 28
            if self.minrange > self.abssmall:
                self.finish=0
                self.noforward=0
                if self.prevlist:
                    self.minrange = self.prevlist.pop()
                self.maxrange = self.minrange+28
                self.count=2
                if (self.minrange) < 2:
                    self.minrange=2
                    self.maxrange=self.maxrange+28
                
                self.frame.destroy()
                self.frame = tk.Frame(self.canvas)
                self.frame.bind("<Configure>", self.onFrameConfigure)
                self.canvas.create_window((4,4), window=self.frame, anchor="nw", 
                                          tags="self.frame")
                self.populate()
        
        def OpenUrl(i, *args):
            webbrowser.open_new(wb['Sheet1'].cell(row=i, column=URLCol).value)
            
        

            
        def loopThruLocations(v, lcol, i, obj, val):

            if obj != None:

                if wb['Sheet1'].cell(row=i, column=lcol).value == 1:
                    
                    obj.array[val] = obj.array[val] + v
                    Scrollbar.setText(obj)
                
                lcol = lcol + 1
                
                if wb['Sheet1'].cell(row=i, column=lcol).value != None:

                    obj = Scrollbar.nameToObj(wb['Sheet1'].cell(row=1, column=lcol).value)
                    loopThruLocations(v, lcol, i, obj, val)
                
            
            
        
        def callback(i, count, *args):
            whatsit = tkvar[count-2].get()
            
            val = wb['Sheet1'].cell(row=i, column=2).value
            if(whatsit=="Unobtained") & (val != 0):
                obj = Scrollbar.nameToObj(wb['Sheet1'].cell(row=i, column=TurnInCol).value)
                alll.array[val]=alll.array[val]-1
                if(wb['Sheet1'].cell(row=i, column=5).value != None):
                    lives.array[val]=lives.array[val]-1
                #subtract from the array
                if val != 1:
                    obj.array[val]=obj.array[val]-1
                elif val == 1:
                    lobj = Scrollbar.nameToObj(wb['Sheet1'].cell(row=1, column=LocationFirstCol).value)
                    loopThruLocations(-1, LocationFirstCol, i, lobj, val)
                    
                
                wb['Sheet1'].cell(row=i, column=2).value=0
                obj.array[0]=obj.array[0]+1
                alll.array[0]=alll.array[0]+1
                if(wb['Sheet1'].cell(row=i, column=5).value != None):
                    lives.array[0]=lives.array[0]+1
                Scrollbar.setText(obj)
                Scrollbar.setText(alll)
                Scrollbar.setText(lives)
                wb.save('FLActive.xlsx')
            elif(whatsit=="Obtained") & (val != 1):
                lobj = Scrollbar.nameToObj(wb['Sheet1'].cell(row=1, column=LocationFirstCol).value)
                alll.array[1]=alll.array[1]+1
                if(wb['Sheet1'].cell(row=i, column=5).value != None):
                    lives.array[1]=lives.array[1]+1
                loopThruLocations(1, LocationFirstCol, i, lobj, 1)
                
                wb['Sheet1'].cell(row=i, column=2).value=1
                obj = Scrollbar.nameToObj(wb['Sheet1'].cell(row=i, column=TurnInCol).value)
                obj.array[val]=obj.array[val]-1
                alll.array[val]=alll.array[val]-1
                if(wb['Sheet1'].cell(row=i, column=5).value != None):
                    lives.array[val]=lives.array[val]-1
                Scrollbar.setText(obj)
                Scrollbar.setText(alll)
                Scrollbar.setText(lives)
                wb.save('FLActive.xlsx')
            elif(whatsit=="Completed") & (val != 2):
                obj = Scrollbar.nameToObj(wb['Sheet1'].cell(row=i, column=TurnInCol).value)
                alll.array[val]=alll.array[val]-1
                if(wb['Sheet1'].cell(row=i, column=5).value != None):
                    lives.array[val]=lives.array[val]-1
                #subtract from the array
                if val != 1:
                    obj.array[val]=obj.array[val]-1
                elif val == 1:
                    lobj = Scrollbar.nameToObj(wb['Sheet1'].cell(row=1, column=LocationFirstCol).value)
                    loopThruLocations(-1, LocationFirstCol, i, lobj, val)
                
                wb['Sheet1'].cell(row=i, column=2).value=2
                obj.array[2]=obj.array[2]+1
                alll.array[2]=alll.array[2]+1
                if(wb['Sheet1'].cell(row=i, column=5).value != None):
                    lives.array[2]=lives.array[2]+1
                Scrollbar.setText(obj)
                Scrollbar.setText(alll)
                Scrollbar.setText(lives)
                wb.save('FLActive.xlsx')
            elif(whatsit=="Turned In") & (val != 3):
                obj = Scrollbar.nameToObj(wb['Sheet1'].cell(row=i, column=TurnInCol).value)
                alll.array[val]=alll.array[val]-1
                if(wb['Sheet1'].cell(row=i, column=5).value != None):
                    lives.array[val]=lives.array[val]-1
                #subtract from the array
                if val != 1:
                    obj.array[val]=obj.array[val]-1
                elif val == 1:
                    lobj = Scrollbar.nameToObj(wb['Sheet1'].cell(row=1, column=LocationFirstCol).value)
                    loopThruLocations(-1, LocationFirstCol, i, lobj, val)
                
                wb['Sheet1'].cell(row=i, column=2).value=3
                obj.array[3]=obj.array[3]+1
                alll.array[3]=alll.array[3]+1
                if(wb['Sheet1'].cell(row=i, column=5).value != None):
                    lives.array[3]=lives.array[3]+1
                Scrollbar.setText(obj)
                Scrollbar.setText(alll)
                Scrollbar.setText(lives)
                wb.save('FLActive.xlsx')
            
        
        choices = ['Unobtained','Obtained','Completed','Turned In']
        for j in range(2,10):
            if (j != URLCol):
                tk.Label(self.frame, text=wb['Sheet1'].cell(row=1, column=j).value).grid(row=1,column=j, sticky='nw')
        tk.Label(self.frame, text="Location").grid(row=1,column=10, sticky='nw')
        
        global LocationObj
        global lives
        global alll
        i=self.minrange
        if LocationObj == lives:
            while self.finish != 1:
                if(i==self.maxrange):
                    self.finish= 1
                for j in range(2,11):
                    if (j==2):
                        var = StringVar()
                        
                        tkvar.append(var)
                        
                        val = wb['Sheet1'].cell(row=i, column=j).value
                        if val == None:
                            self.maxrange=i
                            self.finish = 1
                            break
                        if TopButton == 4:
                            if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                            if wb['Sheet1'].cell(row=i, column=5).value != None:
                                if val == 0:
                                    tkvar[self.count-2].set('Unobtained')
                                elif val == 1:
                                    tkvar[self.count-2].set('Obtained')
                                elif val == 2:
                                    tkvar[self.count-2].set('Completed')
                                elif val == 3:
                                    tkvar[self.count-2].set('Turned In')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            else:
                                self.finish = 1
                                self.noforward = 1
                                break
                        elif TopButton == 3:
                            if (val == 3) & (wb['Sheet1'].cell(row=i, column=5).value != None):
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                tkvar[self.count-2].set('Turned In')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif val < 3:
                                self.finish = 1
                                self.noforward = 1
                                break
                            else:
                                break
                        elif TopButton == 2:
                            if (val == 2) & (wb['Sheet1'].cell(row=i, column=5).value != None):
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                tkvar[self.count-2].set('Completed')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif val > 2:
                                self.minrange=self.minrange+1
                                self.maxrange=self.maxrange+1
                                break
                            elif val < 2:
                                self.finish = 1
                                self.noforward = 1
                                break
                            else:
                                break
                        elif TopButton == 1:
                            if (val == 1) & (wb['Sheet1'].cell(row=i, column=5).value != None):
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                tkvar[self.count-2].set('Obtained')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif val > 1:
                                self.minrange=self.minrange+1
                                self.maxrange=self.maxrange+1
                                break
                            elif val < 1:
                                self.finish = 1
                                self.noforward = 1
                                break
                            else:
                                break
                        elif (TopButton) == 0:
                            if (val == 0) & (wb['Sheet1'].cell(row=i, column=5).value != None):
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                tkvar[self.count-2].set('Unobtained')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif val > 0:
                                self.minrange=self.minrange+1
                                self.maxrange=self.maxrange+1
                                break
                            else:
                                break
                        else:
                            self.finish=1
                            break
                    elif(j==QNameCol):
                        obj=Button(self.frame, text=wb['Sheet1'].cell(row=i, column=QNameCol).value)
                        obj.configure(command=lambda i=i: OpenUrl(i))
                        obj.grid(row=i,column=j, sticky='nw')
                    elif (j==10):
                        locationchoices = []
                        k = 10
                        while wb['Sheet1'].cell(row=1, column=k).value != None:
                            if wb['Sheet1'].cell(row=i, column=k).value == 1:
                                locationchoices.append(wb['Sheet1'].cell(row=1, column=k).value)
                            k=k+1
                        locationvar = StringVar()
                        locationvar.set(locationchoices[0])
                        locationoption = OptionMenu(self.frame, locationvar, *locationchoices)
                        locationoption.grid(row=i, column=j)
                        
                    elif(j!=URLCol):
                        tk.Label(self.frame, text=wb['Sheet1'].cell(row=i, column=j).value).grid(row=i,column=j, sticky='nw')
                i=i+1
                
                  
            backb=Button(self.frame, text="<- Back")
            backb.configure(command=goBack)
            backb.grid(row=i+1,column=4, sticky='nw')
            
            forwardb=Button(self.frame, text="Forward ->")
            forwardb.configure(command=goForward)
            forwardb.grid(row=i+1,column=5, sticky='nw')
        elif LocationObj == alll:
            while self.finish != 1:
                if(i==self.maxrange):
                    self.finish= 1
                for j in range(2,11):
                    if (j==2):
                        var = StringVar()
                        
                        tkvar.append(var)
                        
                        val = wb['Sheet1'].cell(row=i, column=j).value
                        if val == None:
                            self.maxrange=i
                            self.finish = 1
                            break
                        if TopButton == 4:
                            if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                            if val == 0:
                                tkvar[self.count-2].set('Unobtained')
                            elif val == 1:
                                tkvar[self.count-2].set('Obtained')
                            elif val == 2:
                                tkvar[self.count-2].set('Completed')
                            elif val == 3:
                                tkvar[self.count-2].set('Turned In')
                            option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                            option.grid(row=i, column=j)
                            tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                            self.count=self.count+1
                        elif TopButton == 3:
                            if val == 3:
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                tkvar[self.count-2].set('Turned In')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif val < 3:
                                self.finish = 1
                                self.noforward = 1
                                break
                            else:
                                break
                        elif TopButton == 2:
                            if val == 2:
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                tkvar[self.count-2].set('Completed')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif val > 2:
                                self.minrange=self.minrange+1
                                self.maxrange=self.maxrange+1
                                break
                            elif val < 2:
                                self.finish = 1
                                self.noforward = 1
                                break
                            else:
                                break
                        elif TopButton == 1:
                            if val == 1:
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                tkvar[self.count-2].set('Obtained')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif val > 1:
                                self.minrange=self.minrange+1
                                self.maxrange=self.maxrange+1
                                break
                            elif val < 1:
                                self.finish = 1
                                self.noforward = 1
                                break
                            else:
                                break
                        elif TopButton == 0:
                            if val == 0:
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                tkvar[self.count-2].set('Unobtained')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif val > 0:
                                self.minrange=self.minrange+1
                                self.maxrange=self.maxrange+1
                                break
                            else:
                                break
                        else:
                            self.finish=1
                            break
                    elif(j==QNameCol):
                        obj=Button(self.frame, text=wb['Sheet1'].cell(row=i, column=QNameCol).value)
                        obj.configure(command=lambda i=i: OpenUrl(i))
                        obj.grid(row=i,column=j, sticky='nw')
                    elif (j==10):
                        locationchoices = []
                        k = 10
                        while wb['Sheet1'].cell(row=1, column=k).value != None:
                            if wb['Sheet1'].cell(row=i, column=k).value == 1:
                                locationchoices.append(wb['Sheet1'].cell(row=1, column=k).value)
                            k=k+1
                        locationvar = StringVar()
                        locationvar.set(locationchoices[0])
                        locationoption = OptionMenu(self.frame, locationvar, *locationchoices)
                        locationoption.grid(row=i, column=j)
                        
                    elif(j!=URLCol):
                        tk.Label(self.frame, text=wb['Sheet1'].cell(row=i, column=j).value).grid(row=i,column=j, sticky='nw')
                i=i+1
                
                  
            backb=Button(self.frame, text="<- Back")
            backb.configure(command=goBack)
            backb.grid(row=i+1,column=4, sticky='nw')
            
            forwardb=Button(self.frame, text="Forward ->")
            forwardb.configure(command=goForward)
            forwardb.grid(row=i+1,column=5, sticky='nw')
        else:
            haveseen = 0
            while self.finish != 1:
                if(i==1299):
                    self.finish= 1
                for j in range(2,11):
                    if (j==2):
                        
                        
                        val = wb['Sheet1'].cell(row=i, column=j).value
                        if val == None:
                            self.maxrange=i
                            self.finish = 1
                            break
                        if TopButton == 4:
                            if (LocationObj.key == wb['Sheet1'].cell(row=i, column=TurnInCol).value):
                                
                                haveseen = 1
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                var = StringVar()
                            
                                tkvar.append(var)
                                if val == 0:
                                    tkvar[self.count-2].set('Unobtained')
                                elif val == 1:
                                    tkvar[self.count-2].set('Obtained')
                                elif val == 2:
                                    tkvar[self.count-2].set('Completed')
                                elif val == 3:
                                    tkvar[self.count-2].set('Turned In')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif haveseen == 0:
                                self.minrange=self.minrange+1
                                self.maxrange=self.maxrange+1
                                break
                            elif haveseen == 1:
                                self.finish = 1
                                self.noforward = 1
                                break
                        elif TopButton == 3:
                            if (val == 3) & (LocationObj.key == wb['Sheet1'].cell(row=i, column=TurnInCol).value):
                                haveseen = 1
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                var = StringVar()
                        
                                tkvar.append(var)
                                tkvar[self.count-2].set('Turned In')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif (val == 3) & (LocationObj.key != wb['Sheet1'].cell(row=i, column=TurnInCol).value) & (haveseen == 0):
                                self.minrange=self.minrange+1
                                self.maxrange=self.maxrange+1
                                break
                            elif val < 3:
                                self.finish = 1
                                self.noforward = 1
                                break
                            else:
                                break
                        elif TopButton == 2:
                            if (val == 2) & (LocationObj.key == wb['Sheet1'].cell(row=i, column=TurnInCol).value):
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                var = StringVar()
                        
                                tkvar.append(var)
                                tkvar[self.count-2].set('Completed')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif val >= 2:
                                self.minrange=self.minrange+1
                                self.maxrange=self.maxrange+1
                                break
                            elif val < 2:
                                self.finish = 1
                                self.noforward = 1
                                break
                            else:
                                break
                        elif TopButton == 1:
                            if (val == 1) & (Scrollbar.findCol(LocationObj)!= -1):
                                if (wb['Sheet1'].cell(row=i, column=Scrollbar.findCol(LocationObj)).value == 1):
                                    if (self.abssmall == 0) & ((self.count-2) == 0):
                                        self.abssmall = i
                                    var = StringVar()
                            
                                    tkvar.append(var)
                                    tkvar[self.count-2].set('Obtained')
                                    option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                    option.grid(row=i, column=j)
                                    tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                    self.count=self.count+1
                                elif(self.count > 2):
                                    self.finish = 1
                                    self.noforward = 1
                                    break
                            elif val >= 1:
                                self.minrange=self.minrange+1
                                self.maxrange=self.maxrange+1
                                break
                            elif val < 1:
                                self.finish = 1
                                self.noforward = 1
                                break
                            else:
                                break
                        elif TopButton == 0:
                            var = StringVar()
                        
                            tkvar.append(var)
                            if (val == 0) & (LocationObj.key == wb['Sheet1'].cell(row=i, column=TurnInCol).value):
                                if (self.abssmall == 0) & ((self.count-2) == 0):
                                    self.abssmall = i
                                tkvar[self.count-2].set('Unobtained')
                                option = OptionMenu(self.frame, tkvar[self.count-2], *choices)
                                option.grid(row=i, column=j)
                                tkvar[self.count-2].trace("w", lambda a, b, c, i=i, count=self.count: callback(i, count))
                                self.count=self.count+1
                            elif val >= 0:
                                self.minrange=self.minrange+1
                                self.maxrange=self.maxrange+1
                                break
                            else:
                                break
                        else:
                            self.finish=1
                            break
                    elif(j==QNameCol):
                        obj=Button(self.frame, text=wb['Sheet1'].cell(row=i, column=QNameCol).value)
                        obj.configure(command=lambda i=i: OpenUrl(i))
                        obj.grid(row=i,column=j, sticky='nw')
                    elif (j==10):
                        locationchoices = []
                        k = 10
                        while wb['Sheet1'].cell(row=1, column=k).value != None:
                            if wb['Sheet1'].cell(row=i, column=k).value == 1:
                                locationchoices.append(wb['Sheet1'].cell(row=1, column=k).value)
                            k=k+1
                        locationvar = StringVar()
                        locationvar.set(locationchoices[0])
                        locationoption = OptionMenu(self.frame, locationvar, *locationchoices)
                        locationoption.grid(row=i, column=j)
                        
                    elif(j!=URLCol):
                        tk.Label(self.frame, text=wb['Sheet1'].cell(row=i, column=j).value).grid(row=i,column=j, sticky='nw')
                i=i+1
                if self.count == 31:
                    self.finish=1
                
            backb=Button(self.frame, text="<- Back")
            backb.configure(command=goBack)
            backb.grid(row=i+1,column=4, sticky='nw')
            
            forwardb=Button(self.frame, text="Forward ->")
            forwardb.configure(command=goForward)
            forwardb.grid(row=i+1,column=5, sticky='nw')
                    
       
    
    def onFrameConfigure(self, event):
        '''Reset the scroll region to encompass the inner frame'''
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))


class Window(Scrollbar, Location):
    
    

    def __init__(self, master=None):
        Frame.__init__(self, master)               
        self.master = master
        self.init_window()

    def createToolTip(widget, text):
        toolTip = ToolTip(widget)
        def enter(event):
            toolTip.showtip(text)
        def leave(event):
            toolTip.hidetip()
        widget.bind('<Enter>', enter)
        widget.bind('<Leave>', leave)
        
    def setText(obj):
        obj.text.set(str(obj.array[0]) + " / " + str(obj.array[1]) + " / "  + str(obj.array[2]) + " / "  + str(obj.array[3]))
            
    #Creation of init_window
    def init_window(self):
        
        global cc
        cc = Location(self)
        global cs
        cs = Location(self)
        global csd
        csd = Location(self)
        global cad
        cad = Location(self)
        global co
        co = Location(self)
        global sc
        sc = Location(self)
        global wfc
        wfc = Location(self)
        global ec
        ec = Location(self)
        global egp
        egp = Location(self)
        global hc
        hc = Location(self)
        global ms
        ms = Location(self)
        global lc
        lc = Location(self)
        global wc
        wc = Location(self)
        global mss
        mss = Location(self)
        global ew
        ew = Location(self)
        global de
        de = Location(self)
        global ewv
        ewv = Location(self)
        global st
        st = Location(self)
        global wgp
        wgp = Location(self)
        global fp
        fp = Location(self)
        global dr
        dr = Location(self)
        global ppp
        ppp = Location(self)
        global pppw
        pppw = Location(self)
        global ppm
        ppm = Location(self)
        global ppbd
        ppbd = Location(self)
        global ta
        ta = Location(self)
        global nc
        nc = Location(self)
        global dc
        dc = Location(self)
        global dsf
        dsf = Location(self)
        global amsp
        amsp = Location(self)
        global amsa
        amsa = Location(self)
        global amo
        amo = Location(self)
        global ad
        ad = Location(self)
        global cb
        cb = Location(self)
        global sl
        sl = Location(self)
        global ar
        ar = Location(self)
        global cac
        cac = Location(self)
        global cos
        cos = Location(self)
        global pc
        pc = Location(self)
        global tn
        tn = Location(self)
        global cg
        cg = Location(self)
        global fs
        fs = Location(self)
        global rhs
        rhs = Location(self)
        global pb
        pb = Location(self)
        global fg
        fg = Location(self)
        global at
        at = Location(self)
        global alll
        alll = Location(self)
        global lives
        lives  = Location(self)
        
        cc.array=array.array('i',[0,0,0,0])
        cs.array=array.array('i',[0,0,0,0])
        csd.array=array.array('i',[0,0,0,0])
        cad.array=array.array('i',[0,0,0,0])
        co.array=array.array('i',[0,0,0,0])
        sc.array=array.array('i',[0,0,0,0])
        wc.array=array.array('i',[0,0,0,0])
        ec.array=array.array('i',[0,0,0,0])
        egp.array=array.array('i',[0,0,0,0])
        hc.array=array.array('i',[0,0,0,0])
        ms.array=array.array('i',[0,0,0,0])
        lc.array=array.array('i',[0,0,0,0])
        wfc.array=array.array('i',[0,0,0,0])
        mss.array=array.array('i',[0,0,0,0])
        ew.array=array.array('i',[0,0,0,0])
        de.array=array.array('i',[0,0,0,0])
        ewv.array=array.array('i',[0,0,0,0])
        st.array=array.array('i',[0,0,0,0])
        wgp.array=array.array('i',[0,0,0,0])
        fp.array=array.array('i',[0,0,0,0])
        dr.array=array.array('i',[0,0,0,0])
        ppp.array=array.array('i',[0,0,0,0])
        pppw.array=array.array('i',[0,0,0,0])
        ppm.array=array.array('i',[0,0,0,0])
        ppbd.array=array.array('i',[0,0,0,0])
        ta.array=array.array('i',[0,0,0,0])
        nc.array=array.array('i',[0,0,0,0])
        dc.array=array.array('i',[0,0,0,0])
        dsf.array=array.array('i',[0,0,0,0])
        amsp.array=array.array('i',[0,0,0,0])
        amsa.array=array.array('i',[0,0,0,0])
        amo.array=array.array('i',[0,0,0,0])
        ad.array=array.array('i',[0,0,0,0])
        cb.array=array.array('i',[0,0,0,0])
        sl.array=array.array('i',[0,0,0,0])
        ar.array=array.array('i',[0,0,0,0])
        cac.array=array.array('i',[0,0,0,0])
        cos.array=array.array('i',[0,0,0,0])
        pc.array=array.array('i',[0,0,0,0])
        tn.array=array.array('i',[0,0,0,0])
        cg.array=array.array('i',[0,0,0,0])
        fs.array=array.array('i',[0,0,0,0])
        rhs.array=array.array('i',[0,0,0,0])
        pb.array=array.array('i',[0,0,0,0])
        fg.array=array.array('i',[0,0,0,0])
        at.array=array.array('i',[0,0,0,0])
        alll.array=array.array('i',[0,0,0,0])
        lives.array=array.array('i',[0,0,0,0])
        global LocationObj
        LocationObj = alll
        alll.key = "All"
        # changing the title of our master widget    
        def findButtonName():
            global TopButton
            if TopButton == 4:
                return "All Requests"
            elif TopButton == 3:
                return "Turned In Requests"
            elif TopButton == 2:
                return "Completed Requests"
            elif TopButton == 1:
                return "Obtained Requests"
            else:
                return "Unobtained Requests"
        
        self.master.title("Fantasy Life - {} - {}".format(LocationObj.key, findButtonName()))

        # allowing the widget to take the full space of the root window
        self.pack(fill=BOTH, expand=1)


        def starter(i, val, lcol, uppedalll):
            if(val == 1):
                obj=Scrollbar.nameToObj(wb['Sheet1'].cell(row=1, column=lcol).value)
                if obj != None:
                    if wb['Sheet1'].cell(row=i, column=lcol).value == 1:
                        obj.array[val]=obj.array[val]+1
                        if uppedalll == 0:
                            alll.array[val]=alll.array[val]+1
                            if(wb['Sheet1'].cell(row=i, column=5).value != None):
                                lives.array[val]=lives.array[val]+1
                            uppedalll = 1
                    starter(i, val, lcol+1, uppedalll)
            else:
                obj=Scrollbar.nameToObj(wb['Sheet1'].cell(row=i, column=lcol).value)
                if obj != None:
                    alll.array[val]=alll.array[val]+1
                    if(wb['Sheet1'].cell(row=i, column=5).value != None):
                        lives.array[val]=lives.array[val]+1
                    obj.array[val]=obj.array[val]+1
        
        #keep total of each unobtained, obtained, completed, turned in
        i=2
        val = wb['Sheet1'].cell(row=i, column=2).value
        while val != None:
            if val==0: #unobtained
                starter(i,val,TurnInCol, 0)
            elif val==1: #obtained
                starter(i,val,LocationFirstCol, 0)
            elif val==2: #completed
                starter(i,val,TurnInCol, 0)
            elif val==3: #turned in
                starter(i,val,TurnInCol, 0)
            i=i+1
            val = wb['Sheet1'].cell(row=i, column=2).value
        
            
        map_frame = Frame(self)
        map_frame.pack(side = LEFT)
        
        button_frame = Frame(self)
        button_frame.pack()
        scrollframe = Scrollbar(self)
        
        def selectLocation(obj, *args):
            global wb
            global TopButton
            global LocationObj
            LocationObj=obj
            wb.close()
            xl = pd.ExcelFile("FLActive.xlsx")
            df = xl.parse("Sheet1")
            if(TopButton == 1) & (LocationObj != alll):
                try:
                    df = df.sort_values(by=[LocationObj.key, 'Complete', 'NPC', 'Rank'],ascending=False)
                except:
                    df = df.sort_values(by=['Complete', 'NPC', 'Rank'],ascending=False)
            elif(TopButton == 4):
                df = df.sort_values(by=['Turn In', 'NPC', 'Complete', 'Rank'],ascending=False)
            else:
                df = df.sort_values(by=['Complete', 'Turn In', 'NPC', 'Rank'],ascending=False)
            writer = pd.ExcelWriter('FLActive.xlsx')
            df.to_excel(writer,sheet_name='Sheet1',index=False, engine='xlsxwriter')
            writer.save()
            xl.close
            wb = load_workbook('FLActive.xlsx')
            
            
            self.master.title("Fantasy Life - {} - {}".format(LocationObj.key, findButtonName()))
            
            scrollframe.minrange=2
            scrollframe.maxrange=30
            scrollframe.count = scrollframe.minrange
            scrollframe.finish = 0
            scrollframe.noforward = 0
            scrollframe.prevlist = []
            scrollframe.abssmall = 0
            scrollframe.frame.destroy()
            scrollframe.frame = tk.Frame(scrollframe.canvas)
            scrollframe.frame.bind("<Configure>", scrollframe.onFrameConfigure)
            scrollframe.canvas.create_window((4,4), window=scrollframe.frame, anchor="nw", tags="scrollframe.frame")
            scrollframe.populate()
        
        tn.photo = PhotoImage(file="Images/TerraNimbus.gif", master=self)
        tn.label = Label(image=tn.photo, master=self)
        tn.label.image = tn.photo
        Window.setText(tn)
        tn.b = Button(map_frame,textvariable=tn.text,image=tn.photo, command=lambda obj=tn: selectLocation(obj), compound="top")
        tn.b.grid(row=0, column=0)
        tn.key = "Terra Nimbus"
        Window.createToolTip(tn.b, tn.key)
        
        ppp.photo=PhotoImage(file="Images/PortPuertoPalace.gif", master=self)
        ppp.label = Label(image=ppp.photo, master=self)
        ppp.label.image = ppp.photo # keep a reference!
        ppp.text = tk.StringVar()
        Window.setText(ppp)
        ppp.b = Button(map_frame,textvariable=ppp.text,image=ppp.photo, command=lambda obj=ppp: selectLocation(obj), compound="top")
        ppp.b.grid(row=0, column=1)
        ppp.key = "Port Puerto Palace"
        Window.createToolTip(ppp.b, ppp.key)
        
        st.photo=PhotoImage(file="Images/SpiritTree.gif", master=self)
        st.label = Label(image=st.photo, master=self)
        st.label.image = st.photo # keep a reference!
        st.text = tk.StringVar()
        Window.setText(st)
        st.b = Button(map_frame,textvariable=st.text,image=st.photo, command=lambda obj=st: selectLocation(obj), compound="top")
        st.b.grid(row=0, column=2)
        st.key = "Spirit Tree"
        Window.createToolTip(st.b, st.key)
        
        de.photo=PhotoImage(file="Images/DeepElderwood.gif", master=self)
        de.label = Label(image=de.photo, master=self)
        de.label.image = de.photo # keep a reference!
        de.text = tk.StringVar()
        Window.setText(de)
        de.b = Button(map_frame,textvariable=de.text,image=de.photo, command=lambda obj=de: selectLocation(obj), compound="top")
        de.b.grid(row=0, column=3)
        de.key = "Deep Elderwood"
        Window.createToolTip(de.b, de.key)
        
        ms.photo=PhotoImage(file="Images/MountSnowpeak.gif", master=self)
        ms.label = Label(image=ms.photo, master=self)
        ms.label.image = ms.photo # keep a reference!
        ms.text = tk.StringVar()
        Window.setText(ms)
        ms.b = Button(map_frame,textvariable=ms.text,image=ms.photo, command=lambda obj=ms: selectLocation(obj), compound="top")
        ms.b.grid(row=0, column=4)
        ms.key = "Mount Snowpeak"
        Window.createToolTip(ms.b, ms.key)
        
        mss.photo=PhotoImage(file="Images/MountSnowpeakSummit.gif", master=self)
        mss.label = Label(image=mss.photo, master=self)
        mss.label.image = mss.photo # keep a reference!
        mss.text = tk.StringVar()
        Window.setText(mss)
        mss.b = Button(map_frame,textvariable=mss.text,image=mss.photo, command=lambda obj=mss: selectLocation(obj), compound="top")
        mss.b.grid(row=0, column=5)
        mss.key = "Mount Snowpeak Summit"
        Window.createToolTip(mss.b, mss.key)
        
        pc.photo=PhotoImage(file="Images/PlushlingCamp.gif", master=self)
        pc.label = Label(image=pc.photo, master=self)
        pc.label.image = pc.photo # keep a reference!
        pc.text = tk.StringVar()
        Window.setText(pc)
        pc.b = Button(map_frame,textvariable=pc.text,image=pc.photo, command=lambda obj=pc: selectLocation(obj), compound="top")
        pc.b.grid(row=1, column=0)
        pc.key = "Plushling Camp"
        Window.createToolTip(pc.b, pc.key)
        
        pppw.photo=PhotoImage(file="Images/PortPuertoPalaceWay.gif", master=self)
        pppw.label = Label(image=pppw.photo, master=self)
        pppw.label.image = pppw.photo # keep a reference!
        pppw.text = tk.StringVar()
        Window.setText(pppw)
        pppw.b = Button(map_frame,textvariable=pppw.text,image=pppw.photo, command=lambda obj=pppw: selectLocation(obj), compound="top")
        pppw.b.grid(row=1, column=1)
        pppw.key = "Port Puerto Palace Way"
        Window.createToolTip(pppw.b, pppw.key)
        
        ewv.photo=PhotoImage(file="Images/ElderwoodVillage.gif", master=self)
        ewv.label = Label(image=ewv.photo, master=self)
        ewv.label.image = ewv.photo # keep a reference!
        ewv.text = tk.StringVar()
        Window.setText(ewv)
        ewv.b = Button(map_frame,textvariable=ewv.text,image=ewv.photo, command=lambda obj=ewv: selectLocation(obj), compound="top")
        ewv.b.grid(row=1, column=2)
        ewv.key = "Elderwood Village"
        Window.createToolTip(ewv.b, ewv.key)
        
        ew.photo=PhotoImage(file="Images/Elderwood.gif", master=self)
        ew.label = Label(image=ew.photo, master=self)
        ew.label.image = ew.photo # keep a reference!
        ew.text = tk.StringVar()
        Window.setText(ew)
        ew.b = Button(map_frame,textvariable=ew.text,image=ew.photo, command=lambda obj=ew: selectLocation(obj), compound="top")
        ew.b.grid(row=1, column=3)
        ew.key = "Elderwood"
        Window.createToolTip(ew.b, ew.key)
        
        lc.photo=PhotoImage(file="Images/LavaCave.gif", master=self)
        lc.label = Label(image=lc.photo, master=self)
        lc.label.image = lc.photo # keep a reference!
        lc.text = tk.StringVar()
        Window.setText(lc)
        lc.b = Button(map_frame,textvariable=lc.text,image=lc.photo, command=lambda obj=lc: selectLocation(obj), compound="top")
        lc.b.grid(row=1, column=4)
        lc.key = "Lava Cave"
        Window.createToolTip(lc.b, lc.key)
        
        wfc.photo=PhotoImage(file="Images/WaterfallCave.gif", master=self)
        wfc.label = Label(image=wfc.photo, master=self)
        wfc.label.image = wfc.photo # keep a reference!
        wfc.text = tk.StringVar()
        Window.setText(wfc)
        wfc.b = Button(map_frame,textvariable=wfc.text,image=wfc.photo, command=lambda obj=wfc: selectLocation(obj), compound="top")
        wfc.b.grid(row=1, column=5)
        wfc.key = "Waterfall Cave"
        Window.createToolTip(wfc.b, wfc.key)
        
        ta.photo=PhotoImage(file="Images/TortugaArchipelago.gif", master=self)
        ta.label = Label(image=ta.photo, master=self)
        ta.label.image = ta.photo # keep a reference!
        ta.text = tk.StringVar()
        Window.setText(ta)
        ta.b = Button(map_frame,textvariable=ta.text,image=ta.photo, command=lambda obj=ta: selectLocation(obj), compound="top")
        ta.b.grid(row=2, column=0)
        ta.key = "Tortuga Archipelago"
        Window.createToolTip(ta.b, ta.key)
        
        ppm.photo=PhotoImage(file="Images/PortPuertoMarina.gif", master=self)
        ppm.label = Label(image=ppm.photo, master=self)
        ppm.label.image = ppm.photo # keep a reference!
        ppm.text = tk.StringVar()
        Window.setText(ppm)
        ppm.b = Button(map_frame,textvariable=ppm.text,image=ppm.photo, command=lambda obj=ppm: selectLocation(obj), compound="top")
        ppm.b.grid(row=2, column=1)
        ppm.key = "Port Puerto Marina"
        Window.createToolTip(ppm.b, ppm.key)
        
        fp.photo=PhotoImage(file="Images/FarleysPlantation.gif", master=self)
        fp.label = Label(image=fp.photo, master=self)
        fp.label.image = fp.photo # keep a reference!
        fp.text = tk.StringVar()
        Window.setText(fp)
        fp.b = Button(map_frame,textvariable=fp.text,image=fp.photo, command=lambda obj=fp: selectLocation(obj), compound="top")
        fp.b.grid(row=2, column=2)
        fp.key = "Farley's Plantation"
        Window.createToolTip(fp.b, fp.key)
        
        hc.photo=PhotoImage(file="Images/HaniwaCave.gif", master=self)
        hc.label = Label(image=hc.photo, master=self)
        hc.label.image = hc.photo # keep a reference!
        hc.text = tk.StringVar()
        Window.setText(hc)
        hc.b = Button(map_frame,textvariable=hc.text,image=hc.photo, command=lambda obj=hc: selectLocation(obj), compound="top")
        hc.b.grid(row=2, column=3)
        hc.key = "Haniwa Cave"
        Window.createToolTip(hc.b, hc.key)
        
        fs.photo=PhotoImage(file="Images/ForestShrine.gif", master=self)
        fs.label = Label(image=fs.photo, master=self)
        fs.label.image = fs.photo # keep a reference!
        fs.text = tk.StringVar()
        Window.setText(fs)
        fs.b = Button(map_frame,textvariable=fs.text,image=fs.photo, command=lambda obj=fs: selectLocation(obj), compound="top")
        fs.b.grid(row=2, column=4)
        fs.key = "Forest Shrine"
        Window.createToolTip(fs.b, fs.key)
        
        at.photo=PhotoImage(file="Images/AncientTower.gif", master=self)
        at.label = Label(image=at.photo, master=self)
        at.label.image = at.photo # keep a reference!
        at.text = tk.StringVar()
        Window.setText(at)
        at.b = Button(map_frame,textvariable=at.text,image=at.photo, command=lambda obj=at: selectLocation(obj), compound="top")
        at.b.grid(row=2, column=5)
        at.key = "Ancient Tower"
        Window.createToolTip(at.b, at.key)
        
        nc.photo=PhotoImage(file="Images/NautilusCave.gif", master=self)
        nc.label = Label(image=nc.photo, master=self)
        nc.label.image = nc.photo # keep a reference!
        nc.text = tk.StringVar()
        Window.setText(nc)
        nc.b = Button(map_frame,textvariable=nc.text,image=nc.photo, command=lambda obj=nc: selectLocation(obj), compound="top")
        nc.b.grid(row=3, column=0)
        nc.key = "Nautilus Cave"
        Window.createToolTip(nc.b, nc.key)
        
        ppbd.photo=PhotoImage(file="Images/PortPuertoBeachDistrict.gif", master=self)
        ppbd.label = Label(image=ppbd.photo, master=self)
        ppbd.label.image = ppbd.photo # keep a reference!
        ppbd.text = tk.StringVar()
        Window.setText(ppbd)
        ppbd.b = Button(map_frame,textvariable=ppbd.text,image=ppbd.photo, command=lambda obj=ppbd: selectLocation(obj), compound="top")
        ppbd.b.grid(row=3, column=1)
        ppbd.key = "Port Puerto Beach District"
        Window.createToolTip(ppbd.b, ppbd.key)
        
        wgp.photo=PhotoImage(file="Images/WestGrassyPlains.gif", master=self)
        wgp.label = Label(image=wgp.photo, master=self)
        wgp.label.image = wgp.photo # keep a reference!
        wgp.text = tk.StringVar()
        Window.setText(wgp)
        wgp.b = Button(map_frame,textvariable=wgp.text,image=wgp.photo, command=lambda obj=wgp: selectLocation(obj), compound="top")
        wgp.b.grid(row=3, column=2)
        wgp.key = "West Grassy Plains"
        Window.createToolTip(wgp.b, wgp.key)
        
        egp.photo=PhotoImage(file="Images/EastGrassyPlains.gif", master=self)
        egp.label = Label(image=egp.photo, master=self)
        egp.label.image = egp.photo # keep a reference!
        egp.text = tk.StringVar()
        Window.setText(egp)
        egp.b = Button(map_frame,textvariable=egp.text,image=egp.photo, command=lambda obj=egp: selectLocation(obj), compound="top")
        egp.b.grid(row=3, column=3)
        egp.key = "East Grassy Plains"
        Window.createToolTip(egp.b, egp.key)
        
        rhs.photo=PhotoImage(file="Images/RockyHillShrine.gif", master=self)
        rhs.label = Label(image=rhs.photo, master=self)
        rhs.label.image = rhs.photo # keep a reference!
        rhs.text = tk.StringVar()
        Window.setText(rhs)
        rhs.b = Button(map_frame,textvariable=rhs.text,image=rhs.photo, command=lambda obj=rhs: selectLocation(obj), compound="top")
        rhs.b.grid(row=3, column=4)
        rhs.key = "Rocky Hill Shrine"
        Window.createToolTip(rhs.b, rhs.key)
        
        cg.photo=PhotoImage(file="Images/CentralGrassland.gif", master=self)
        cg.label = Label(image=cg.photo, master=self)
        cg.label.image = cg.photo # keep a reference!
        cg.text = tk.StringVar()
        Window.setText(cg)
        cg.b = Button(map_frame,textvariable=cg.text,image=cg.photo, command=lambda obj=cg: selectLocation(obj), compound="top")
        cg.b.grid(row=3, column=5)
        cg.key = "Central Grassland"
        Window.createToolTip(cg.b, cg.key)
           
        dc.photo=PhotoImage(file="Images/DeepseaCave.gif", master=self)
        dc.label = Label(image=dc.photo, master=self)
        dc.label.image = dc.photo # keep a reference!
        dc.text = tk.StringVar()
        Window.setText(dc)
        dc.b = Button(map_frame,textvariable=dc.text,image=dc.photo, command=lambda obj=dc: selectLocation(obj), compound="top")
        dc.b.grid(row=4, column=0)
        dc.key = "Deepsea Cave"
        Window.createToolTip(dc.b, dc.key)
        
        dr.photo=PhotoImage(file="Images/DesertRavine.gif", master=self)
        dr.label = Label(image=dr.photo, master=self)
        dr.label.image = dr.photo # keep a reference!
        dr.text = tk.StringVar()
        Window.setText(dr)
        dr.b = Button(map_frame,textvariable=dr.text,image=dr.photo, command=lambda obj=dr: selectLocation(obj), compound="top")
        dr.b.grid(row=4, column=1)
        dr.key = "Desert Ravine"
        Window.createToolTip(dr.b, dr.key)
        
        cc.photo=PhotoImage(file="Images/CasteleCastle.gif", master=self)
        cc.label = Label(image=cc.photo, master=self)
        cc.label.image = cc.photo # keep a reference!
        cc.text = tk.StringVar()
        Window.setText(cc)
        cc.b = Button(map_frame,textvariable=cc.text,image=cc.photo, command=lambda obj=cc: selectLocation(obj), compound="top")
        cc.b.grid(row=4, column=3)
        cc.key = "Castele Castle"
        Window.createToolTip(cc.b, cc.key)
        
        pb.photo=PhotoImage(file="Images/PenguinBeach.gif", master=self)
        pb.label = Label(image=pb.photo, master=self)
        pb.label.image = pb.photo # keep a reference!
        pb.text = tk.StringVar()
        Window.setText(pb)
        pb.b = Button(map_frame,textvariable=pb.text,image=pb.photo, command=lambda obj=pb: selectLocation(obj), compound="top")
        pb.b.grid(row=4, column=4)
        pb.key = "Penguin Beach"
        Window.createToolTip(pb.b, pb.key)
        
        fg.photo=PhotoImage(file="Images/FurlinsGrotto.gif", master=self)
        fg.label = Label(image=fg.photo, master=self)
        fg.label.image = fg.photo # keep a reference!
        fg.text = tk.StringVar()
        Window.setText(fg)
        fg.b = Button(map_frame,textvariable=fg.text,image=fg.photo, command=lambda obj=fg: selectLocation(obj), compound="top")
        fg.b.grid(row=4, column=5)
        fg.key = "Furlin's Grotto"
        Window.createToolTip(fg.b, fg.key)
    
        dsf.photo=PhotoImage(file="Images/DarkSultansFortress.gif", master=self)
        dsf.label = Label(image=dsf.photo, master=self)
        dsf.label.image = dsf.photo # keep a reference!
        dsf.text = tk.StringVar()
        Window.setText(dsf)
        dsf.b = Button(map_frame,textvariable=dsf.text,image=dsf.photo, command=lambda obj=dsf: selectLocation(obj), compound="top")
        dsf.b.grid(row=5, column=0)
        dsf.key = "Dark Sultan's Fortress"
        Window.createToolTip(dsf.b, dsf.key)
        
        wc.photo=PhotoImage(file="Images/WestCastele.gif", master=self)
        wc.label = Label(image=wc.photo, master=self)
        wc.label.image = wc.photo # keep a reference!
        wc.text = tk.StringVar()
        Window.setText(wc)
        wc.b = Button(map_frame,textvariable=wc.text,image=wc.photo, command=lambda obj=wc: selectLocation(obj), compound="top")
        wc.b.grid(row=5, column=1)
        wc.key = "West Castele"
        Window.createToolTip(wc.b, wc.key)
        
        csd.photo=PhotoImage(file="Images/CasteleShoppingDistrict.gif", master=self)
        csd.label = Label(image=csd.photo, master=self)
        csd.label.image = csd.photo # keep a reference!
        csd.text = tk.StringVar()
        Window.setText(csd)
        csd.b = Button(map_frame,textvariable=csd.text,image=csd.photo, command=lambda obj=csd: selectLocation(obj), compound="top")
        csd.b.grid(row=5, column=2)
        csd.key = "Castele Shopping District"
        Window.createToolTip(csd.b, csd.key)
        
        cs.photo=PhotoImage(file="Images/CasteleSquare.gif", master=self)
        cs.label = Label(image=cs.photo, master=self)
        cs.label.image = cs.photo # keep a reference!
        cs.text = tk.StringVar()
        Window.setText(cs)
        cs.b = Button(map_frame,textvariable=cs.text,image=cs.photo, command=lambda obj=cs: selectLocation(obj), compound="top")
        cs.b.grid(row=5, column=3)
        cs.key = "Castele Square"
        Window.createToolTip(cs.b, cs.key)
        
        cad.photo=PhotoImage(file="Images/CasteleArtisanDistrict.gif", master=self)
        cad.label = Label(image=cad.photo, master=self)
        cad.label.image = cad.photo # keep a reference!
        cad.text = tk.StringVar()
        Window.setText(cad)
        cad.b = Button(map_frame,textvariable=cad.text,image=cad.photo, command=lambda obj=cad: selectLocation(obj), compound="top")
        cad.b.grid(row=5, column=4)
        cad.key = "Castele Artisan District"
        Window.createToolTip(cad.b, cad.key)
        
        ec.photo=PhotoImage(file="Images/EastCastele.gif", master=self)
        ec.label = Label(image=ec.photo, master=self)
        ec.label.image = ec.photo # keep a reference!
        ec.text = tk.StringVar()
        Window.setText(ec)
        ec.b = Button(map_frame,textvariable=ec.text,image=ec.photo, command=lambda obj=ec: selectLocation(obj), compound="top")
        ec.b.grid(row=5, column=5)
        ec.key = "East Castele"
        Window.createToolTip(ec.b, ec.key)
        
        amsp.photo=PhotoImage(file="Images/AlMaajikSpelltown.gif", master=self)
        amsp.label = Label(image=amsp.photo, master=self)
        amsp.label.image = amsp.photo # keep a reference!
        amsp.text = tk.StringVar()
        Window.setText(amsp)
        amsp.b = Button(map_frame,textvariable=amsp.text,image=amsp.photo, command=lambda obj=amsp: selectLocation(obj), compound="top")
        amsp.b.grid(row=6, column=0)
        amsp.key = "Al Maajik Spelltown"
        Window.createToolTip(amsp.b, amsp.key)
        
        amsa.photo=PhotoImage(file="Images/AlMaajikSandtown.gif", master=self)
        amsa.label = Label(image=amsa.photo, master=self)
        amsa.label.image = amsa.photo # keep a reference!
        amsa.text = tk.StringVar()
        Window.setText(amsa)
        amsa.b = Button(map_frame,textvariable=amsa.text,image=amsa.photo, command=lambda obj=amsa: selectLocation(obj), compound="top")
        amsa.b.grid(row=6, column=1)
        amsa.key = "Al Maajik Sandtown"
        Window.createToolTip(amsa.b, amsa.key)
        
        amo.photo=PhotoImage(file="Images/AlMaajikOutskirts.gif", master=self)
        amo.label = Label(image=amo.photo, master=self)
        amo.label.image = amo.photo # keep a reference!
        amo.text = tk.StringVar()
        Window.setText(amo)
        amo.b = Button(map_frame,textvariable=amo.text,image=amo.photo, command=lambda obj=amo: selectLocation(obj), compound="top")
        amo.b.grid(row=6, column=2)
        amo.key = "Al Maajik Outskirts"
        Window.createToolTip(amo.b, amo.key)
        
        sc.photo=PhotoImage(file="Images/SouthCastele.gif", master=self)
        sc.label = Label(image=sc.photo, master=self)
        sc.label.image = sc.photo # keep a reference!
        sc.text = tk.StringVar()
        Window.setText(sc)
        sc.b = Button(map_frame,textvariable=sc.text,image=sc.photo, command=lambda obj=sc: selectLocation(obj), compound="top")
        sc.b.grid(row=6, column=3)
        sc.key = "South Castele"
        Window.createToolTip(sc.b, sc.key)
        
        cb.photo=PhotoImage(file="Images/CaveofBones.gif", master=self)
        cb.label = Label(image=cb.photo, master=self)
        cb.label.image = cb.photo # keep a reference!
        cb.text = tk.StringVar()
        Window.setText(cb)
        cb.b = Button(map_frame,textvariable=cb.text,image=cb.photo, command=lambda obj=cb: selectLocation(obj), compound="top")
        cb.b.grid(row=6, column=4)
        cb.key = "Cave of Bones"
        Window.createToolTip(cb.b, cb.key)
        
        ar.photo=PhotoImage(file="Images/AncientRuins.gif", master=self)
        ar.label = Label(image=ar.photo, master=self)
        ar.label.image = ar.photo # keep a reference!
        ar.text = tk.StringVar()
        Window.setText(ar)
        ar.b = Button(map_frame,textvariable=ar.text,image=ar.photo, command=lambda obj=ar: selectLocation(obj), compound="top")
        ar.b.grid(row=6, column=5)
        ar.key = "Ancient Ruins"
        Window.createToolTip(ar.b, ar.key)
        
        cac.photo=PhotoImage(file="Images/CactoCove.gif", master=self)
        cac.label = Label(image=cac.photo, master=self)
        cac.label.image = cac.photo # keep a reference!
        cac.text = tk.StringVar()
        Window.setText(cac)
        cac.b = Button(map_frame,textvariable=cac.text,image=cac.photo, command=lambda obj=cac: selectLocation(obj), compound="top")
        cac.b.grid(row=7, column=0)
        cac.key = "Cacto Cove"
        Window.createToolTip(cac.b, cac.key)
        
        cos.photo=PhotoImage(file="Images/CaveofShadows.gif", master=self)
        cos.label = Label(image=cos.photo, master=self)
        cos.label.image = cos.photo # keep a reference!
        cos.text = tk.StringVar()
        Window.setText(cos)
        cos.b = Button(map_frame,textvariable=cos.text,image=cos.photo, command=lambda obj=cos: selectLocation(obj), compound="top")
        cos.b.grid(row=7, column=1)
        cos.key = "Cave of Shadows"
        Window.createToolTip(cos.b, cos.key)
        
        ad.photo=PhotoImage(file="Images/AridianDesert.gif", master=self)
        ad.label = Label(image=ad.photo, master=self)
        ad.label.image = ad.photo # keep a reference!
        ad.text = tk.StringVar()
        Window.setText(ad)
        ad.b = Button(map_frame,textvariable=ad.text,image=ad.photo, command=lambda obj=ad: selectLocation(obj), compound="top")
        ad.b.grid(row=7, column=2)
        ad.key = "Aridian Desert"
        Window.createToolTip(ad.b, ad.key)
        
        co.photo=PhotoImage(file="Images/CasteleOutskirts.gif", master=self)
        co.label = Label(image=co.photo, master=self)
        co.label.image = co.photo # keep a reference!
        co.text = tk.StringVar()
        Window.setText(co)
        co.b = Button(map_frame,textvariable=co.text,image=co.photo, command=lambda obj=co: selectLocation(obj), compound="top")
        co.b.grid(row=7, column=3)
        co.key = "Castele Outskirts"
        Window.createToolTip(co.b, co.key)
        
        sl.photo=PhotoImage(file="Images/SubterraneanLake.gif", master=self)
        sl.label = Label(image=sl.photo, master=self)
        sl.label.image = sl.photo # keep a reference!
        sl.text = tk.StringVar()
        Window.setText(sl)
        sl.b = Button(map_frame,textvariable=sl.text,image=sl.photo, command=lambda obj=sl: selectLocation(obj), compound="top")
        sl.b.grid(row=7, column=4)
        sl.key = "Subterranean Lake"
        Window.createToolTip(sl.b, sl.key)
        
        alll.photo=PhotoImage(file="Images/all.gif", master=self)
        alll.label = Label(image=alll.photo, master=self)
        alll.label.image = alll.photo # keep a reference!
        alll.text = tk.StringVar()
        Window.setText(alll)
        alll.b = Button(map_frame,textvariable=alll.text,image=alll.photo, command=lambda obj=alll: selectLocation(obj), compound="top")
        alll.b.grid(row=7, column=5)
        Window.createToolTip(alll.b, alll.key)
        
        
        
        def lifeB(obj, *args):
            global wb
            global TopButton
            global LocationObj
            LocationObj=obj
            wb.close()
            xl = pd.ExcelFile("FLActive.xlsx")
            df = xl.parse("Sheet1")

            if(TopButton == 4):
                df = df.sort_values(by=['Life', 'Rank', 'Complete'],ascending=False)
            else:
                df = df.sort_values(by=['Complete', 'Life', 'Rank'],ascending=False)
            writer = pd.ExcelWriter('FLActive.xlsx')
            df.to_excel(writer,sheet_name='Sheet1',index=False, engine='xlsxwriter')
            writer.save()
            xl.close
            wb = load_workbook('FLActive.xlsx')
            
            self.master.title("Fantasy Life - {} - {}".format(LocationObj.key, findButtonName()))
            
            scrollframe.minrange=2
            scrollframe.maxrange=30
            scrollframe.count = scrollframe.minrange
            scrollframe.finish = 0
            scrollframe.noforward = 0
            scrollframe.prevlist = []
            scrollframe.abssmall = 0
            scrollframe.frame.destroy()
            scrollframe.frame = tk.Frame(scrollframe.canvas)
            scrollframe.frame.bind("<Configure>", scrollframe.onFrameConfigure)
            scrollframe.canvas.create_window((4,4), window=scrollframe.frame, anchor="nw", tags="scrollframe.frame")
            scrollframe.populate()
        
        
        lives.photo=PhotoImage(file="Images/lives.gif", master=self)
        lives.label = Label(image=lives.photo, master=self)
        lives.label.image = lives.photo # keep a reference!
        lives.text = tk.StringVar()
        Window.setText(lives)
        lives.b = Button(map_frame,textvariable=lives.text,image=lives.photo, command=lambda obj=lives: lifeB(obj), compound="top")
        lives.b.grid(row=4, column=2)
        lives.key = "Lives"
        Window.createToolTip(lives.b, lives.key)
        
        
        
        
        def topB(v, *args):
            global wb
            global TopButton
            global LocationObj
            TopButton = v
            wb.close()
            xl = pd.ExcelFile("FLActive.xlsx")
            df = xl.parse("Sheet1")
            if(TopButton == 1) & (LocationObj != alll):
                if (LocationObj == lives):
                    df = df.sort_values(by=['Complete', 'Life', 'Rank'],ascending=False)
                elif Scrollbar.findCol(LocationObj)!= -1:
                    try:
                        df = df.sort_values(by=[LocationObj.key, 'Complete', 'NPC', 'Rank'],ascending=False)
                    except:
                        df = df.sort_values(by=['Complete', 'NPC', 'Rank'],ascending=False)
            elif(TopButton == 4):
                if(LocationObj == lives):
                    df = df.sort_values(by=['Life', 'Rank', 'Complete'],ascending=False)
                else:
                    df = df.sort_values(by=['Turn In', 'NPC', 'Complete', 'Rank'],ascending=False)
            else:
                if (LocationObj == lives):
                    df = df.sort_values(by=['Complete', 'Life', 'Rank'],ascending=False)
                else:
                    df = df.sort_values(by=['Complete', 'Turn In', 'NPC', 'Rank'],ascending=False)
            writer = pd.ExcelWriter('FLActive.xlsx')
            df.to_excel(writer,sheet_name='Sheet1',index=False, engine='xlsxwriter')
            writer.save()
            xl.close
            wb = load_workbook('FLActive.xlsx')
            
            self.master.title("Fantasy Life - {} - {}".format(LocationObj.key, findButtonName()))
            
            scrollframe.minrange=2
            scrollframe.maxrange=30
            scrollframe.count = scrollframe.minrange
            scrollframe.finish = 0
            scrollframe.noforward = 0
            scrollframe.prevlist = []
            scrollframe.abssmall = 0
            scrollframe.frame.destroy()
            scrollframe.frame = tk.Frame(scrollframe.canvas)
            scrollframe.frame.bind("<Configure>", scrollframe.onFrameConfigure)
            scrollframe.canvas.create_window((4,4), window=scrollframe.frame, anchor="nw", tags="scrollframe.frame")
            scrollframe.populate()
            
            
            
        
        
        # creating buttons
        unobtainedBut = Button(button_frame, text="Unobtained Requests", command=lambda v=0: topB(v))
        obtainedBut = Button(button_frame, text="Obtained Requests", command=lambda v=1: topB(v))
        completedBut = Button(button_frame, text="Completed Requests", command=lambda v=2: topB(v))
        turnedinBut = Button(button_frame, text="Turned In Requests", command=lambda v=3: topB(v))
        allBut = Button(button_frame, text="All Requests", command=lambda v=4: topB(v))
        
        # placing buttons
        unobtainedBut.grid(row=0, column=0)
        obtainedBut.grid(row=0, column=1)
        completedBut.grid(row=0, column=2)
        turnedinBut.grid(row=0, column=3)
        allBut.grid(row=0, column=4)
        

        
        scrollframe.pack(side="top", fill="both", expand=True)
        
   
    
    def client_exit(self):
        exit() 

try:
    #xl = pd.ExcelFile('FLActive.xlsx')
    global wb
    wb = load_workbook('FLActive.xlsx')
    # Store configuration file values
except FileNotFoundError:
    try:
        xld = pd.ExcelFile('FLData.xlsx')
        copyfile('FLData.xlsx', 'FLActive.xlsx')
        #xl = pd.ExcelFile('FLActive.xlsx')
        wb = load_workbook('FLActive.xlsx')
        xld.close()
        
    except FileNotFoundError:
        print("ERROR")
        sys.exit()
      
        

    



root = Tk()

#size of the window
root.geometry("1500x1000")
root.iconbitmap(r'Images/icon.ico')


app = Window(root)


root.mainloop()
wb.close()
